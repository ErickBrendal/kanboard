#!/usr/bin/env python3
"""
Importa as 14 demandas faltantes diretamente do Excel
"""

import requests
import json
import openpyxl
from datetime import datetime
import time

BASE_URL = "http://kanboard.eblsolucoescorp.tec.br"
API_URL = f"{BASE_URL}/jsonrpc.php"
API_TOKEN = "446685558386787e2b1e0cebb0141257bebea7b2496bc6e960559c950692"
EXCEL_PATH = "/home/ubuntu/upload/Base_Fast_Tracking_Outubro.xlsx"
PROJECT_ID = 11

session = requests.Session()
session.auth = ("admin", API_TOKEN)
session.headers.update({"Content-Type": "application/json"})

_req_id = 0

def api(method, params=None):
    global _req_id
    _req_id += 1
    payload = {"jsonrpc": "2.0", "method": method, "id": _req_id, "params": params or {}}
    try:
        r = session.post(API_URL, json=payload, timeout=30)
        data = r.json()
        if "error" in data:
            print(f"  API ERROR {method}: {data['error']}")
            return None
        return data.get("result")
    except Exception as e:
        print(f"  EXCEPTION {method}: {e}")
        return None

FASE_MAP = {
    "backlog/sem priorização": 165,
    "backlog": 165,
    "sem priorização": 165,
    "pendente": 165,
    "refinamento": 166,
    " refinamento": 166,
    "priorizada": 167,
    "em análise": 168,
    "análise": 168,
    "em estimativa": 169,
    "estimativa": 169,
    "aguardando aprovação": 170,
    "em aprovação": 170,
    "aprovação": 170,
    "em desenvolvimento": 171,
    "desenvolvimento": 171,
    "em atendimento": 171,
    "atendimento": 171,
    "em homologação": 172,
    "homologação": 172,
    "homogação": 172,
    "deploy": 173,
    "em implementação": 173,
    "implementação": 173,
    "implementado": 174,
    "concluído": 174,
    "cancelado": 175,
}

AREA_COLORS = {
    "Ar & Eletro": "orange",
    "Automação": "purple",
    "Bens de consumo": "blue",
    "Cadastros": "yellow",
    "Comercial": "green",
    "Corporativo": "grey",
    "Diretoria": "red",
    "Financeiro": "lime",
    "Fiscal": "teal",
    "Logística": "cyan",
    "Marketing-Refrigeração": "pink",
    "Programação de Materiais": "brown",
    "Refrigeração": "sky",
    "TI": "navy",
}

PRIORIDADE_MAP = {1: 3, 2: 2, 3: 1, 4: 0, "1": 3, "2": 2, "3": 1, "4": 0}

TIPO_MAP = {
    "Parametrização": "Parametrização",
    "Produtividade Comercial": "Produtividade Comercial",
    "Produtividade TI": "Produtividade TI",
    "Incidente ": "Incidente",
    "Incidente": "Incidente",
}

def get_existing_titles():
    tasks = api("getAllTasks", {"project_id": PROJECT_ID, "status_id": 1}) or []
    return set(t["title"] for t in tasks)

def get_categories():
    cats = api("getAllCategories", {"project_id": PROJECT_ID}) or []
    return {c["name"]: c["id"] for c in cats}

def get_swimlanes():
    lanes = api("getAllSwimlanes", {"project_id": PROJECT_ID}) or []
    return {s["name"]: s["id"] for s in lanes}

def format_date(dt):
    if not dt:
        return None
    if isinstance(dt, datetime):
        return int(dt.timestamp())
    return None

def main():
    print("=" * 60)
    print("IMPORTAÇÃO DAS DEMANDAS FALTANTES")
    print("=" * 60)
    
    me = api("getMe")
    if not me:
        print("ERRO: Falha na autenticação!")
        return
    print(f"✅ Conectado como: {me.get('name')}")
    
    existing = get_existing_titles()
    cats = get_categories()
    lanes = get_swimlanes()
    
    print(f"Tarefas existentes: {len(existing)}")
    
    # Carregar Excel
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Status Report']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    
    created = 0
    skipped = 0
    errors = 0
    
    for row in range(2, ws.max_row+1):
        d = {}
        for c in range(1, ws.max_column+1):
            h = headers[c-1]
            if h:
                d[str(h).strip()] = ws.cell(row, c).value
        
        topico = str(d.get('Tópico') or '').strip()
        if not topico or topico == 'None':
            continue
        
        # Verificar se já existe (comparação exata)
        if topico in existing:
            skipped += 1
            continue
        
        # Fase -> Coluna
        fase_raw = str(d.get('Fase Atual') or '').strip()
        col_id = FASE_MAP.get(fase_raw.lower(), 165)
        
        # Tipo -> Categoria
        tipo_raw = str(d.get('Tipo de Demanda') or '').strip()
        cat_name = TIPO_MAP.get(tipo_raw, "Melhoria")
        cat_id = cats.get(cat_name)
        
        # Responsável -> Swimlane
        resp = str(d.get('Responsavel pela demanda') or '').strip()
        lane_id = lanes.get(resp)
        
        # Prioridade
        prio_raw = d.get('Prioridade area ') or d.get('Prioridade') or 3
        priority = PRIORIDADE_MAP.get(prio_raw, 1)
        
        # Área e cor
        area = str(d.get('Área Solicitante') or '').strip()
        color = AREA_COLORS.get(area, "blue")
        
        # Datas
        due_date = format_date(d.get('Go Live'))
        
        # Campos adicionais
        cherwell_id = d.get('ID - Cherwell') or ''
        rdm = d.get('N°RDM') or ''
        valtech = d.get('N° Valtech') or ''
        horas_v = d.get('Horas estimadas Valtech') or ''
        horas_e = d.get('Horas estimadas Elgin') or ''
        valor = d.get('Valor') or ''
        requisitante = d.get('Requisitante') or ''
        aprovado = d.get('Aprovado ?') or ''
        data_aprov = d.get('Data da Aprovação') or ''
        aprovado_por = d.get('Aprovado por:') or ''
        obs = d.get('Obs:') or ''
        seq = d.get('Sequencia Demanda') or ''
        previsao = d.get('Previsão Etapa') or ''
        
        desc = f"""## {topico}

---

### 📋 Identificação
| Campo | Valor |
|-------|-------|
| **ID Cherwell** | {cherwell_id} |
| **Sequência** | {seq} |
| **N° RDM** | {rdm} |
| **N° Valtech** | {valtech} |
| **Fase Atual** | {fase_raw} |
| **Tipo** | {tipo_raw} |

### 👥 Responsabilidade
| Campo | Valor |
|-------|-------|
| **Área Solicitante** | {area} |
| **Responsável** | {resp} |
| **Requisitante** | {requisitante} |

### ⏱️ Estimativas e Financeiro
| Campo | Valor |
|-------|-------|
| **Horas Valtech** | {horas_v}h |
| **Horas Elgin** | {horas_e}h |
| **Valor Estimado** | R$ {valor} |
| **Previsão Etapa** | {previsao} |

### ✅ Aprovação
| Campo | Valor |
|-------|-------|
| **Status** | {aprovado} |
| **Data** | {data_aprov} |
| **Aprovado por** | {aprovado_por} |

### 📝 Observações
{obs if obs else 'Sem observações.'}
"""
        
        task_params = {
            "project_id": PROJECT_ID,
            "title": topico,
            "description": desc.strip(),
            "priority": priority,
            "color_id": color,
            "column_id": col_id,
        }
        
        if cat_id:
            task_params["category_id"] = cat_id
        if lane_id:
            task_params["swimlane_id"] = lane_id
        if due_date:
            task_params["date_due"] = due_date
        
        task_id = api("createTask", task_params)
        
        if task_id:
            created += 1
            existing.add(topico)
            print(f"  ✅ Criada: {topico[:60]}")
        else:
            errors += 1
            print(f"  ❌ Erro: {topico[:60]}")
        
        time.sleep(0.1)
    
    print(f"\n{'='*60}")
    print(f"✅ Criadas: {created}")
    print(f"⏭️  Ignoradas: {skipped}")
    print(f"❌ Erros: {errors}")
    
    # Total final
    all_tasks = api("getAllTasks", {"project_id": PROJECT_ID, "status_id": 1}) or []
    print(f"\n📊 Total final: {len(all_tasks)} tarefas")
    
    col_names = {
        165: "01. Backlog", 166: "02. Refinamento", 167: "03. Priorizada",
        168: "04. Análise", 169: "05. Estimativa", 170: "06. Aprovação",
        171: "07. Desenvolvimento", 172: "08. Homologação", 173: "09. Deploy",
        174: "10. Implementado", 175: "11. Cancelado",
    }
    
    col_dist = {}
    for t in all_tasks:
        cid = t.get('column_id')
        col_name = col_names.get(cid, f"Col {cid}")
        col_dist[col_name] = col_dist.get(col_name, 0) + 1
    
    print("\n📊 Distribuição por fase:")
    for col, cnt in sorted(col_dist.items()):
        bar = "█" * cnt
        print(f"  {col:25s}: {cnt:3d} {bar}")

if __name__ == "__main__":
    main()
