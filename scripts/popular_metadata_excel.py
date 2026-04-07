#!/usr/bin/env python3
"""
EBL Kanboard — Popular Metadados do Excel nas Tarefas
Versão: 1.0 | 2026-04-07

Lê o arquivo PortifólioTI-Elgin2.xlsx e popula os metadados (via metaMagik)
em cada tarefa do Kanboard, fazendo match pelo título da tarefa.
"""

import requests
import json
import time
import openpyxl
from pathlib import Path
from datetime import datetime, date

# ===== CONFIGURAÇÕES =====
TOKEN      = "ea99d4c7d96dbad1b1a1defd79f92286884e1902015ff96731ce624e6317"
URL        = "http://kanboard.eblsolucoescorp.tec.br/jsonrpc.php"
AUTH       = ("jsonrpc", TOKEN)
PROJECT_ID = 1

EXCEL_PATH = Path("/home/ubuntu/upload/PortifólioTI-Elgin2.xlsx")
BASE_DIR   = Path(__file__).parent.parent

_req_id = 0

def log(msg, level="INFO"):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] [{level}] {msg}")

def api(method, params=None):
    global _req_id
    _req_id += 1
    payload = {
        "jsonrpc": "2.0",
        "method": method,
        "id": _req_id,
        "params": params or {}
    }
    r = requests.post(URL, json=payload, auth=AUTH, timeout=15)
    resp = r.json()
    if "error" in resp:
        return None
    return resp.get("result")

def normalize_title(title):
    """Normaliza título para comparação fuzzy"""
    if not title:
        return ""
    return str(title).strip().lower().replace("  ", " ")

def format_date(val):
    """Formata data para string"""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()

def format_value(val):
    """Formata valor genérico para string"""
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    if isinstance(val, str) and val.startswith("="):
        return ""  # Fórmula Excel - ignorar
    return str(val).strip()

def load_excel_data():
    """Carrega dados do Excel e retorna lista de dicts"""
    log(f"Carregando Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Mapear cabeçalhos
    headers_raw = [str(cell.value) if cell.value else "" for cell in ws[1]]

    # Mapeamento de colunas Excel → chaves de metadados
    COL_MAP = {
        "Área TI":                                  "area_ti",
        "Subcategoria":                             "subcategoria",
        "Responsável TI":                           "responsavel_ti",
        "Área Negócio":                             "area_negocio",
        "Fase":                                     "fase_excel",
        "Cherwell":                                 "cherwell_id",
        "Projeto":                                  "titulo_excel",
        "Classificação Risco":                      "classificacao_risco",
        "Data GoLive":                              "data_golive_original",
        "Data GoLive\n(nova data estimada)":        "data_golive_estimada",
        "Data início Projeto\nPlanejado":           "data_inicio_planejado",
        "Data Fim Projeto\nPlanejado":              "data_fim_planejado",
        "% de Evolução \nProjeto  Planejado\n":     "evolucao_planejado_pct",
        "% de Evolução \nProjeto Realizado\n":      "evolucao_realizado_pct",
        "Status (Desvio)\n%":                       "desvio_pct",
        "Recurso Interno ou Externo\n(Identificação Custo)": "recurso_tipo",
        "Custo Planejado":                          "custo_planejado",
        "Custo Realizado":                          "custo_realizado",
        "ROI Aprovado":                             "roi_aprovado",
        "Valor ROI \n(Se aplicado)":                "valor_roi",
    }

    # Normalizar mapeamento (remover espaços extras)
    col_map_norm = {}
    for k, v in COL_MAP.items():
        col_map_norm[k.strip().replace("\n", "\n")] = v

    # Mapear índices de colunas
    col_indices = {}
    for i, h in enumerate(headers_raw):
        h_clean = h.strip()
        for key, meta_key in col_map_norm.items():
            if h_clean == key.strip() or normalize_title(h_clean) == normalize_title(key):
                col_indices[i] = meta_key
                break

    log(f"Colunas mapeadas: {len(col_indices)}/{len(headers_raw)}")

    # Carregar linhas
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        record = {}
        for i, val in enumerate(row):
            if i in col_indices:
                key = col_indices[i]
                if key in ("data_golive_original", "data_golive_estimada",
                           "data_inicio_planejado", "data_fim_planejado"):
                    record[key] = format_date(val)
                elif key in ("evolucao_planejado_pct", "evolucao_realizado_pct",
                             "desvio_pct", "custo_planejado", "custo_realizado", "valor_roi"):
                    v = format_value(val)
                    record[key] = v if v and not v.startswith("=") else "0"
                else:
                    record[key] = format_value(val)

        if record.get("titulo_excel"):
            rows.append(record)

    log(f"Registros carregados do Excel: {len(rows)}")
    return rows

def get_all_tasks():
    """Obtém todas as tarefas do projeto"""
    tasks = api("getAllTasks", {"project_id": PROJECT_ID, "status_id": 1}) or []
    closed = api("getAllTasks", {"project_id": PROJECT_ID, "status_id": 0}) or []
    all_tasks = tasks + closed
    log(f"Tarefas no Kanboard: {len(all_tasks)} (abertas: {len(tasks)}, fechadas: {len(closed)})")
    return all_tasks

def match_task(excel_title, tasks):
    """Encontra a tarefa no Kanboard pelo título"""
    excel_norm = normalize_title(excel_title)

    # Match exato
    for t in tasks:
        if normalize_title(t.get("title", "")) == excel_norm:
            return t

    # Match parcial (título do Excel contido no título da tarefa ou vice-versa)
    for t in tasks:
        task_norm = normalize_title(t.get("title", ""))
        if excel_norm in task_norm or task_norm in excel_norm:
            return t

    # Match por Cherwell ID (se disponível no título da tarefa)
    return None

def main():
    log("=" * 60)
    log("EBL Kanboard — Popular Metadados do Excel")
    log("=" * 60)

    # Verificar conexão
    version = api("getVersion")
    if not version:
        log("Falha na conexão!", "ERROR")
        return 1
    log(f"Kanboard: {version}")

    # Carregar dados do Excel
    excel_rows = load_excel_data()

    # Obter tarefas do Kanboard
    tasks = get_all_tasks()
    task_by_id = {t["id"]: t for t in tasks}

    # Estatísticas
    matched = 0
    not_matched = []
    updated = 0
    errors = 0

    log("")
    log("Populando metadados...")

    for row in excel_rows:
        titulo = row.get("titulo_excel", "")
        if not titulo:
            continue

        # Encontrar tarefa correspondente
        task = match_task(titulo, tasks)

        if not task:
            not_matched.append(titulo)
            continue

        matched += 1
        task_id = task["id"]

        # Preparar metadados (remover campos vazios)
        metadata = {k: v for k, v in row.items() if v and v != "0" and k != "titulo_excel"}

        # Adicionar timestamp de atualização
        metadata["ultima_atualizacao_excel"] = datetime.now().strftime("%Y-%m-%d %H:%M")

        # Salvar metadados via metaMagik
        result = api("saveTaskMetadata", {"task_id": task_id, "values": metadata})

        if result:
            updated += 1
            log(f"  ✓ #{task_id} '{titulo[:50]}' → {len(metadata)} campos")
        else:
            errors += 1
            log(f"  ✗ #{task_id} '{titulo[:50]}' → Falha ao salvar", "WARN")

        time.sleep(0.2)

    # Resumo
    log("")
    log("=" * 60)
    log("RESULTADO")
    log("=" * 60)
    log(f"  Registros no Excel:     {len(excel_rows)}")
    log(f"  Tarefas encontradas:    {matched}")
    log(f"  Metadados atualizados:  {updated}")
    log(f"  Erros:                  {errors}")
    log(f"  Não encontradas:        {len(not_matched)}")

    if not_matched:
        log("")
        log("Tarefas do Excel não encontradas no Kanboard:")
        for t in not_matched[:10]:
            log(f"  - {t[:70]}")
        if len(not_matched) > 10:
            log(f"  ... e mais {len(not_matched) - 10}")

    # Salvar resultado
    result_data = {
        "timestamp": datetime.now().isoformat(),
        "excel_registros": len(excel_rows),
        "tarefas_encontradas": matched,
        "metadados_atualizados": updated,
        "erros": errors,
        "nao_encontradas": not_matched
    }
    result_path = BASE_DIR / "powerbi" / "metadata_result.json"
    with open(result_path, "w", encoding="utf-8") as f:
        json.dump(result_data, f, ensure_ascii=False, indent=2)
    log(f"\nResultado salvo em: {result_path}")

    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())
