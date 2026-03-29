#!/usr/bin/env python3
"""
Script de Reimportação Completa - Remove duplicatas e importa todas as 105 demandas
"""

import requests
import json
import openpyxl
from datetime import datetime
import time
import sys

BASE_URL = "http://kanboard.eblsolucoescorp.tec.br"
API_URL = f"{BASE_URL}/jsonrpc.php"
API_TOKEN = "446685558386787e2b1e0cebb0141257bebea7b2496bc6e960559c950692"
EXCEL_PATH = "/home/ubuntu/upload/Base_Fast_Tracking_Outubro.xlsx"
PROJECT_ID = 11  # [SF] Fast Track — Salesforce

session = requests.Session()
session.auth = ("admin", API_TOKEN)
session.headers.update({"Content-Type": "application/json"})

_req_id = 0

def api(method, params=None):
    global _req_id
    _req_id += 1
    payload = {"jsonrpc": "2.0", "method": method, "id": _req_id, "params": params or {}}
    try:
        r = session.post(API_URL, json=payload, timeout=30)
        data = r.json()
        if "error" in data:
            return None
        return data.get("result")
    except Exception as e:
        return None

# ============================================================
# MAPEAMENTOS
# ============================================================

FASE_MAP = {
    "backlog/sem priorização": "01. Backlog",
    "backlog": "01. Backlog",
    "sem priorização": "01. Backlog",
    "pendente": "01. Backlog",
    "refinamento": "02. Refinamento",
    " refinamento": "02. Refinamento",
    "priorizada": "03. Priorizada",
    "em análise": "04. Análise",
    "análise": "04. Análise",
    "em estimativa": "05. Estimativa",
    "estimativa": "05. Estimativa",
    "aguardando aprovação": "06. Aprovação",
    "em aprovação": "06. Aprovação",
    "aprovação": "06. Aprovação",
    "em desenvolvimento": "07. Desenvolvimento",
    "desenvolvimento": "07. Desenvolvimento",
    "em atendimento": "07. Desenvolvimento",
    "atendimento": "07. Desenvolvimento",
    "em homologação": "08. Homologação",
    "homologação": "08. Homologação",
    "homogação": "08. Homologação",
    "deploy": "09. Deploy",
    "em implementação": "09. Deploy",
    "implementação": "09. Deploy",
    "implementado": "10. Implementado",
    "concluído": "10. Implementado",
    "cancelado": "11. Cancelado",
}

TIPO_MAP = {
    "Parametrização": "Parametrização",
    "Produtividade Comercial": "Produtividade Comercial",
    "Produtividade TI": "Produtividade TI",
    "Incidente ": "Incidente",
    "Incidente": "Incidente",
}

AREA_COLORS = {
    "Ar & Eletro": "orange",
    "Automação": "purple",
    "Bens de consumo": "blue",
    "Cadastros": "yellow",
    "Comercial": "green",
    "Corporativo": "grey",
    "Diretoria": "red",
    "Financeiro": "lime",
    "Fiscal": "teal",
    "Logística": "cyan",
    "Marketing-Refrigeração": "pink",
    "Programação de Materiais": "brown",
    "Refrigeração": "sky",
    "TI": "navy",
    "Tecnologia da Informação": "navy",
}

PRIORIDADE_MAP = {1: 3, 2: 2, 3: 1, 4: 0, "1": 3, "2": 2, "3": 1, "4": 0}

def get_columns():
    cols = api("getColumns", {"project_id": PROJECT_ID}) or []
    return {c["title"]: c["id"] for c in cols}

def get_swimlanes():
    lanes = api("getAllSwimlanes", {"project_id": PROJECT_ID}) or []
    return {s["name"]: s["id"] for s in lanes}

def get_categories():
    cats = api("getAllCategories", {"project_id": PROJECT_ID}) or []
    return {c["name"]: c["id"] for c in cats}

def get_all_tasks():
    """Obtém todas as tarefas (ativas e concluídas)"""
    active = api("getAllTasks", {"project_id": PROJECT_ID, "status_id": 1}) or []
    closed = api("getAllTasks", {"project_id": PROJECT_ID, "status_id": 0}) or []
    return active + closed

def delete_all_tasks():
    """Remove todas as tarefas do projeto para reimportação limpa"""
    print("Removendo todas as tarefas existentes...")
    tasks = get_all_tasks()
    print(f"  Total a remover: {len(tasks)}")
    removed = 0
    for t in tasks:
        result = api("removeTask", {"task_id": t["id"]})
        if result:
            removed += 1
        time.sleep(0.05)
    print(f"  Removidas: {removed}")
    return removed

def load_excel():
    """Carrega dados do Status Report"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Status Report']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    
    demandas = []
    for row in range(2, ws.max_row+1):
        d = {}
        for c in range(1, ws.max_column+1):
            h = headers[c-1]
            if h:
                d[str(h).strip()] = ws.cell(row, c).value
        
        # Pular linhas sem dados essenciais
        topico = d.get('Tópico') or d.get('Titulo da Demanda')
        if not topico or str(topico).strip() in ('', 'None'):
            continue
        
        demandas.append(d)
    
    return demandas

def format_date(dt):
    if not dt:
        return None
    if isinstance(dt, datetime):
        return int(dt.timestamp())
    return None

def import_all(demandas, cols, lanes, cats):
    """Importa todas as demandas"""
    created = 0
    errors = 0
    error_list = []
    
    for i, d in enumerate(demandas, 1):
        topico = str(d.get('Tópico') or d.get('Titulo da Demanda') or '').strip()
        if not topico:
            continue
        
        # Fase -> Coluna
        fase_raw = str(d.get('Fase Atual') or '').strip()
        col_title = FASE_MAP.get(fase_raw.lower(), "01. Backlog")
        col_id = cols.get(col_title)
        
        # Tipo -> Categoria
        tipo_raw = str(d.get('Tipo de Demanda') or '').strip()
        cat_name = TIPO_MAP.get(tipo_raw, "Melhoria")
        cat_id = cats.get(cat_name)
        
        # Responsável -> Swimlane
        resp = str(d.get('Responsavel pela demanda') or '').strip()
        lane_id = lanes.get(resp)
        
        # Prioridade
        prio_raw = d.get('Prioridade area ') or d.get('Prioridade') or 3
        priority = PRIORIDADE_MAP.get(prio_raw, 1)
        
        # Área e cor
        area = str(d.get('Área Solicitante') or '').strip()
        color = AREA_COLORS.get(area, "blue")
        
        # Datas
        due_date = format_date(d.get('Go Live'))
        
        # Campos adicionais
        cherwell_id = d.get('ID - Cherwell') or ''
        rdm = d.get('N°RDM') or ''
        valtech = d.get('N° Valtech') or ''
        horas_v = d.get('Horas estimadas Valtech') or ''
        horas_e = d.get('Horas estimadas Elgin') or ''
        valor = d.get('Valor') or ''
        requisitante = d.get('Requisitante') or ''
        aprovado = d.get('Aprovado ?') or ''
        data_aprov = d.get('Data da Aprovação') or ''
        aprovado_por = d.get('Aprovado por:') or ''
        obs = d.get('Obs:') or ''
        seq = d.get('Sequencia Demanda') or i
        previsao = d.get('Previsão Etapa') or ''
        dev = d.get('Desenvolvimento:') or ''
        
        # Descrição rica em Markdown
        desc = f"""## Demanda #{seq} — {topico}

---

### 📋 Identificação
| Campo | Valor |
|-------|-------|
| **ID Cherwell** | {cherwell_id} |
| **N° RDM** | {rdm} |
| **N° Valtech** | {valtech} |
| **Fase Atual** | {fase_raw} |
| **Tipo** | {tipo_raw} |

### 👥 Responsabilidade
| Campo | Valor |
|-------|-------|
| **Área Solicitante** | {area} |
| **Responsável** | {resp} |
| **Requisitante** | {requisitante} |

### ⏱️ Estimativas
| Campo | Valor |
|-------|-------|
| **Horas Valtech** | {horas_v}h |
| **Horas Elgin** | {horas_e}h |
| **Valor Estimado** | R$ {valor} |
| **Previsão Etapa** | {previsao} |
| **Desenvolvimento** | {dev} |

### ✅ Aprovação
| Campo | Valor |
|-------|-------|
| **Status** | {aprovado} |
| **Data** | {data_aprov} |
| **Aprovado por** | {aprovado_por} |

### 📝 Observações
{obs}
"""
        
        # Montar parâmetros da tarefa
        task_params = {
            "project_id": PROJECT_ID,
            "title": topico,
            "description": desc.strip(),
            "priority": priority,
            "color_id": color,
        }
        
        if col_id:
            task_params["column_id"] = col_id
        if cat_id:
            task_params["category_id"] = cat_id
        if lane_id:
            task_params["swimlane_id"] = lane_id
        if due_date:
            task_params["date_due"] = due_date
        
        task_id = api("createTask", task_params)
        
        if task_id:
            created += 1
            if created % 10 == 0:
                print(f"  {created}/{len(demandas)} importadas...")
        else:
            errors += 1
            error_list.append(topico[:60])
        
        time.sleep(0.08)
    
    return created, errors, error_list

def main():
    print("=" * 60)
    print("REIMPORTAÇÃO COMPLETA - KANBOARD EBL")
    print("=" * 60)
    
    # Verificar conexão
    me = api("getMe")
    if not me:
        print("ERRO: Falha na autenticação!")
        sys.exit(1)
    print(f"✅ Conectado como: {me.get('name')}")
    
    # Verificar projeto
    proj = api("getProjectById", {"project_id": PROJECT_ID})
    if not proj:
        print(f"ERRO: Projeto ID {PROJECT_ID} não encontrado!")
        sys.exit(1)
    print(f"✅ Projeto: {proj['name']}")
    
    # Remover todas as tarefas existentes
    print("\n[1/4] Limpando tarefas existentes...")
    delete_all_tasks()
    
    # Carregar dados do Excel
    print("\n[2/4] Carregando dados do Excel...")
    demandas = load_excel()
    print(f"✅ {len(demandas)} demandas encontradas")
    
    # Obter estrutura do projeto
    print("\n[3/4] Obtendo estrutura do projeto...")
    cols = get_columns()
    lanes = get_swimlanes()
    cats = get_categories()
    print(f"  Colunas: {list(cols.keys())}")
    print(f"  Swimlanes: {list(lanes.keys())}")
    print(f"  Categorias: {list(cats.keys())}")
    
    # Importar todas as demandas
    print("\n[4/4] Importando demandas...")
    created, errors, error_list = import_all(demandas, cols, lanes, cats)
    
    print(f"\n{'='*60}")
    print(f"✅ IMPORTAÇÃO CONCLUÍDA!")
    print(f"  Criadas: {created}")
    print(f"  Erros: {errors}")
    if error_list:
        print(f"  Erros em: {error_list[:5]}")
    print(f"{'='*60}")
    
    # Verificar resultado
    tasks = get_all_tasks()
    print(f"\n📊 Total de tarefas no projeto: {len(tasks)}")
    
    # Contar por coluna
    col_count = {}
    for t in tasks:
        col_name = t.get('column_name', 'Sem coluna')
        col_count[col_name] = col_count.get(col_name, 0) + 1
    
    print("\n📊 Distribuição por fase:")
    for col, cnt in sorted(col_count.items()):
        print(f"  {col}: {cnt} tarefas")

if __name__ == "__main__":
    main()
