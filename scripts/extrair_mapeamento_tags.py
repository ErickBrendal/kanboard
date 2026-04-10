import openpyxl
import json

wb = openpyxl.load_workbook('/home/ubuntu/upload/PortifólioTI-Elgin2.xlsx', data_only=True)
ws = wb.active

# Mapeamentos de normalização
AREA_TI_MAP = {
    'CRM Salesforce': '[TI] CRM Salesforce',
    'Dados & Digital': '[TI] Dados & Digital',
}

SUBCATEGORIA_MAP = {
    'Suporte Salesforce': '[SUB] Suporte Salesforce',
    'CRM Salesforce':     '[SUB] CRM Salesforce',
    'Integração':         '[SUB] Integração',
    'E-Commerce':         '[SUB] E-Commerce',
    'Digital':            '[SUB] Digital',
    'Site e Portal de Apoio': '[SUB] Site e Portal',
    'Dados e BI':         '[SUB] Dados e BI',
}

AREA_NEGOCIO_MAP = {
    'Financeiro':              '[NEG] Financeiro',
    'eCommerce':               '[NEG] eCommerce',
    'E-Commerce':              '[NEG] eCommerce',
    'Comercial':               '[NEG] Comercial',
    'Logistica':               '[NEG] Logística',
    'Logística':               '[NEG] Logística',
    'Fiscal':                  '[NEG] Fiscal',
    'Controladoria':           '[NEG] Controladoria',
    'Marketing':               '[NEG] Marketing',
    'TI':                      '[NEG] TI',
    'Cadastros':               '[NEG] Cadastros',
    'SAC':                     '[NEG] SAC',
    'DHO':                     '[NEG] DHO',
    'Diretoria':               '[NEG] Diretoria',
    'Ar & Eletro':             '[NEG] Ar & Eletro',
    'Refrigeração':            '[NEG] Refrigeração',
    'Automação':               '[NEG] Automação',
    'Bens de Consumo':         '[NEG] Bens de Consumo',
    'Bens de consumo':         '[NEG] Bens de Consumo',
    'Engenharia':              '[NEG] Engenharia',
    'Programação de Materiais':'[NEG] Prog. Materiais',
    'Produção':                '[NEG] Engenharia',
}

RISCO_MAP = {
    'Baixo': '[RISCO] Baixo',
    'Médio': '[RISCO] Médio',
    'Alto':  '[RISCO] Alto',
}

RECURSO_MAP = {
    'Interno': '[REC] Interno',
    'Externo': '[REC] Externo',
    'Ambos':   '[REC] Ambos',
}

ROI_MAP = {
    'Qualitativo':        '[ROI] Qualitativo',
    'Legal / Mandatório': '[ROI] Legal / Mandatório',
    'N/A':                '[ROI] N/A',
}

def evolucao_tag(pct_realizado):
    """Converte % realizado (0.0 a 1.0) em tag de faixa."""
    if pct_realizado is None:
        return None
    try:
        v = float(pct_realizado)
    except:
        return None
    if v == 0:
        return '[EVO] 0%'
    elif v <= 0.25:
        return '[EVO] 1–25%'
    elif v <= 0.50:
        return '[EVO] 26–50%'
    elif v <= 0.75:
        return '[EVO] 51–75%'
    elif v < 1.0:
        return '[EVO] 76–99%'
    else:
        return '[EVO] 100% Concluído'

# Processar planilha
mapeamento = {}  # cherwell_id -> {titulo, tags}
sem_cherwell = []

for row in range(2, ws.max_row + 1):
    area_ti      = str(ws.cell(row, 1).value or '').strip()
    subcategoria = str(ws.cell(row, 2).value or '').strip()
    responsavel  = str(ws.cell(row, 3).value or '').strip()
    area_negocio = str(ws.cell(row, 4).value or '').strip()
    fase         = str(ws.cell(row, 5).value or '').strip()
    cherwell     = str(ws.cell(row, 6).value or '').strip()
    projeto      = str(ws.cell(row, 7).value or '').strip()
    risco        = str(ws.cell(row, 8).value or '').strip()
    pct_realizado = ws.cell(row, 14).value
    recurso      = str(ws.cell(row, 16).value or '').strip()
    roi          = str(ws.cell(row, 19).value or '').strip()

    if not cherwell or not projeto:
        continue

    tags = []

    # Montar tags
    if area_ti in AREA_TI_MAP:
        tags.append(AREA_TI_MAP[area_ti])
    if subcategoria in SUBCATEGORIA_MAP:
        tags.append(SUBCATEGORIA_MAP[subcategoria])
    if area_negocio in AREA_NEGOCIO_MAP:
        tags.append(AREA_NEGOCIO_MAP[area_negocio])
    if risco in RISCO_MAP:
        tags.append(RISCO_MAP[risco])
    if recurso in RECURSO_MAP:
        tags.append(RECURSO_MAP[recurso])
    if roi and roi in ROI_MAP:
        tags.append(ROI_MAP[roi])

    # Tag de evolução (só para fases em andamento)
    fases_andamento = ['Em atendimento', 'Homologação', 'Pendente-Aguardando']
    if any(f in fase for f in fases_andamento):
        evo = evolucao_tag(pct_realizado)
        if evo:
            tags.append(evo)

    mapeamento[cherwell] = {
        'projeto': projeto[:80],  # truncar para legibilidade
        'responsavel': responsavel,
        'fase': fase,
        'tags': tags
    }

# Salvar JSON
with open('/home/ubuntu/kanboard/docs/mapeamento_tags.json', 'w', encoding='utf-8') as f:
    json.dump(mapeamento, f, ensure_ascii=False, indent=2)

print(f"Total de demandas mapeadas: {len(mapeamento)}")
print(f"\nExemplos:")
for k, v in list(mapeamento.items())[:5]:
    print(f"\n  Cherwell: {k}")
    print(f"  Projeto:  {v['projeto'][:60]}")
    print(f"  Tags:     {v['tags']}")
