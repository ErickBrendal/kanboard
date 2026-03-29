#!/usr/bin/env python3
"""
Setup INSANO do Kanboard - EBL Soluções Corporativas
Reestrutura boards, sanitiza colunas, importa todas as demandas do Excel
"""

import requests
import json
import openpyxl
import time
import re
from datetime import datetime

BASE_URL = "http://kanboard.eblsolucoescorp.tec.br/jsonrpc.php"
AUTH = ("admin", "Senha@2026")
EXCEL_PATH = "/home/ubuntu/upload/Base_Fast_Tracking_Outubro.xlsx"

def api(method, params={}, retries=3):
    for i in range(retries):
        try:
            r = requests.post(BASE_URL, auth=AUTH, 
                json={"jsonrpc":"2.0","method":method,"id":1,"params":params}, 
                timeout=20)
            result = r.json()
            if "result" in result:
                return result["result"]
            elif "error" in result:
                print(f"  ⚠️  API Error [{method}]: {result['error']}")
                return None
        except Exception as e:
            if i < retries - 1:
                time.sleep(2)
            else:
                print(f"  ❌ Exception [{method}]: {e}")
    return None

def sanitize_column_name(name):
    """Remove prefixos 'Em' e padroniza nomes de colunas"""
    replacements = {
        "Em Análise": "Análise",
        "Em Estimativa": "Estimativa",
        "Em Aprovação": "Aprovação",
        "Em Desenvolvimento": "Desenvolvimento",
        "Em Homologação": "Homologação",
        "Em Implementação": "Implementação",
        "Em Deploy": "Deploy",
        "Em Produção": "Produção",
        "Em Revisão": "Revisão",
        "Em Teste": "Teste",
        "Em Progresso": "Em Progresso",
    }
    for old, new in replacements.items():
        name = name.replace(old, new)
    return name

# Estrutura padrão de colunas para todos os boards
COLUNAS_PADRAO = [
    {"title": "01. Backlog",        "task_limit": 0, "color": "#95a5a6", "description": "Demandas aguardando priorização"},
    {"title": "02. Refinamento",    "task_limit": 0, "color": "#3498db", "description": "Detalhamento e escopo da demanda"},
    {"title": "03. Priorizada",     "task_limit": 0, "color": "#e67e22", "description": "Aprovada para execução"},
    {"title": "04. Análise",        "task_limit": 0, "color": "#9b59b6", "description": "Análise técnica e funcional"},
    {"title": "05. Estimativa",     "task_limit": 0, "color": "#1abc9c", "description": "Estimativa de esforço e prazo"},
    {"title": "06. Aprovação",      "task_limit": 0, "color": "#f39c12", "description": "Aguardando aprovação do cliente"},
    {"title": "07. Desenvolvimento","task_limit": 0, "color": "#2980b9", "description": "Em desenvolvimento"},
    {"title": "08. Homologação",    "task_limit": 0, "color": "#8e44ad", "description": "Teste e validação pelo cliente"},
    {"title": "09. Deploy",         "task_limit": 0, "color": "#27ae60", "description": "Publicação em produção"},
    {"title": "10. Implementado",   "task_limit": 0, "color": "#2ecc71", "description": "Concluído e entregue"},
    {"title": "11. Cancelado",      "task_limit": 0, "color": "#e74c3c", "description": "Cancelado ou descartado"},
]

# Swimlanes padrão
SWIMLANES_PADRAO = [
    {"name": "🔴 Alta Prioridade",   "description": "Demandas críticas e urgentes"},
    {"name": "🟡 Média Prioridade",  "description": "Demandas importantes"},
    {"name": "🟢 Baixa Prioridade",  "description": "Melhorias e otimizações"},
    {"name": "📋 Backlog Geral",     "description": "Demandas sem prioridade definida"},
]

# Categorias padrão com cores
CATEGORIAS_PADRAO = [
    {"name": "Bug / Correção",       "color_id": "red"},
    {"name": "Nova Funcionalidade",  "color_id": "blue"},
    {"name": "Melhoria",             "color_id": "green"},
    {"name": "Integração",           "color_id": "purple"},
    {"name": "Relatório / BI",       "color_id": "orange"},
    {"name": "Configuração",         "color_id": "yellow"},
    {"name": "Suporte",              "color_id": "grey"},
]

# Mapeamento de status do Excel para colunas do Kanboard
STATUS_MAP = {
    "Backlog":              "01. Backlog",
    "Refinamento":          "02. Refinamento",
    "Priorizada":           "03. Priorizada",
    "Análise":              "04. Análise",
    "Em Análise":           "04. Análise",
    "Estimativa":           "05. Estimativa",
    "Em Estimativa":        "05. Estimativa",
    "Aprovação":            "06. Aprovação",
    "Em Aprovação":         "06. Aprovação",
    "Aguardando Aprovação": "06. Aprovação",
    "Desenvolvimento":      "07. Desenvolvimento",
    "Em Desenvolvimento":   "07. Desenvolvimento",
    "Homologação":          "08. Homologação",
    "Em Homologação":       "08. Homologação",
    "Deploy":               "09. Deploy",
    "Em Deploy":            "09. Deploy",
    "Implementado":         "10. Implementado",
    "Concluído":            "10. Implementado",
    "Cancelado":            "11. Cancelado",
    "Cancelada":            "11. Cancelado",
}

# Mapeamento de prioridade
PRIORITY_MAP = {
    "Alta":    3,
    "Média":   2,
    "Baixa":   1,
    "Crítica": 3,
    "Urgente": 3,
}

# Cores por prioridade
COLOR_MAP = {
    3: "red",
    2: "orange",
    1: "green",
    0: "blue",
}

def get_or_create_user(username, name, email, role="app-user"):
    """Busca ou cria usuário"""
    users = api("getAllUsers") or []
    for u in users:
        if u.get("username") == username:
            return u["id"]
    
    result = api("createUser", {
        "username": username,
        "name": name,
        "email": email,
        "password": "EBL@2026",
        "role": role
    })
    return result if result else None

def setup_project_columns(project_id, project_name):
    """Configura colunas padronizadas em um projeto"""
    print(f"\n  📋 Configurando colunas do projeto {project_name}...")
    
    # Buscar colunas existentes
    existing_cols = api("getColumns", {"project_id": project_id}) or []
    existing_titles = {c["title"]: c["id"] for c in existing_cols}
    
    # Renomear colunas com "Em" 
    for col in existing_cols:
        sanitized = sanitize_column_name(col["title"])
        if sanitized != col["title"]:
            api("updateColumn", {
                "column_id": col["id"],
                "title": sanitized,
                "task_limit": 0
            })
            print(f"    ✅ Renomeado: '{col['title']}' → '{sanitized}'")
    
    # Recarregar colunas após renomeação
    existing_cols = api("getColumns", {"project_id": project_id}) or []
    existing_titles = {c["title"]: c["id"] for c in existing_cols}
    
    # Criar colunas que não existem
    for col_def in COLUNAS_PADRAO:
        if col_def["title"] not in existing_titles:
            result = api("addColumn", {
                "project_id": project_id,
                "title": col_def["title"],
                "task_limit": col_def["task_limit"]
            })
            if result:
                print(f"    ➕ Criada coluna: {col_def['title']}")
    
    # Recarregar e retornar mapeamento
    final_cols = api("getColumns", {"project_id": project_id}) or []
    return {c["title"]: c["id"] for c in final_cols}

def setup_project_swimlanes(project_id, project_name):
    """Configura swimlanes em um projeto"""
    print(f"  🏊 Configurando swimlanes do projeto {project_name}...")
    
    existing_sw = api("getAllSwimlanes", {"project_id": project_id}) or []
    existing_names = {s["name"]: s["id"] for s in existing_sw}
    
    for sw_def in SWIMLANES_PADRAO:
        if sw_def["name"] not in existing_names:
            result = api("addSwimlane", {
                "project_id": project_id,
                "name": sw_def["name"],
                "description": sw_def["description"]
            })
            if result:
                print(f"    ➕ Criada swimlane: {sw_def['name']}")
    
    final_sw = api("getAllSwimlanes", {"project_id": project_id}) or []
    return {s["name"]: s["id"] for s in final_sw}

def setup_project_categories(project_id, project_name):
    """Configura categorias em um projeto"""
    print(f"  🏷️  Configurando categorias do projeto {project_name}...")
    
    existing_cats = api("getAllCategories", {"project_id": project_id}) or []
    existing_names = {c["name"]: c["id"] for c in existing_cats}
    
    for cat_def in CATEGORIAS_PADRAO:
        if cat_def["name"] not in existing_names:
            result = api("createCategory", {
                "project_id": project_id,
                "name": cat_def["name"]
            })
            if result:
                print(f"    ➕ Criada categoria: {cat_def['name']}")
    
    final_cats = api("getAllCategories", {"project_id": project_id}) or []
    return {c["name"]: c["id"] for c in final_cats}

def read_excel_data():
    """Lê os dados do Excel e retorna lista de demandas"""
    print("\n📊 Lendo dados do Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Tentar aba principal
    sheet_names = wb.sheetnames
    print(f"  Abas encontradas: {sheet_names}")
    
    # Usar a primeira aba ou a que tem mais dados
    ws = None
    for sname in ["Cherwell", "Base", "Demandas", "Sheet1", sheet_names[0]]:
        if sname in sheet_names:
            ws = wb[sname]
            print(f"  Usando aba: {sname}")
            break
    
    if not ws:
        ws = wb[sheet_names[0]]
    
    # Ler cabeçalhos
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")
    
    print(f"  Colunas encontradas: {headers}")
    
    # Mapear índices das colunas relevantes
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if any(x in h_lower for x in ["número", "numero", "id", "chamado", "ticket", "incidente"]):
            col_map["id"] = i
        elif any(x in h_lower for x in ["título", "titulo", "assunto", "descrição resumida", "resumo", "demanda"]):
            col_map["title"] = i
        elif any(x in h_lower for x in ["status", "fase", "etapa", "situação"]):
            col_map["status"] = i
        elif any(x in h_lower for x in ["responsável", "responsavel", "analista", "assignee", "atribuído"]):
            col_map["assignee"] = i
        elif any(x in h_lower for x in ["prioridade", "priority", "urgência"]):
            col_map["priority"] = i
        elif any(x in h_lower for x in ["área", "area", "departamento", "setor", "solicitante"]):
            col_map["area"] = i
        elif any(x in h_lower for x in ["data abertura", "criação", "criado", "aberto", "data criação"]):
            col_map["created"] = i
        elif any(x in h_lower for x in ["prazo", "data prevista", "previsão", "entrega", "deadline"]):
            col_map["due_date"] = i
        elif any(x in h_lower for x in ["descrição", "descricao", "detalhes", "observação"]):
            col_map["description"] = i
        elif any(x in h_lower for x in ["categoria", "tipo", "classificação"]):
            col_map["category"] = i
        elif any(x in h_lower for x in ["complexidade", "esforço", "esforco", "pontos"]):
            col_map["complexity"] = i
        elif any(x in h_lower for x in ["sprint", "iteração", "release"]):
            col_map["sprint"] = i
    
    print(f"  Mapeamento de colunas: {col_map}")
    
    # Ler dados
    tasks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Linha vazia
            continue
        
        def get_val(key):
            if key in col_map:
                val = row[col_map[key]]
                return str(val).strip() if val is not None else ""
            return ""
        
        title = get_val("title")
        if not title or title == "None":
            continue
        
        task = {
            "excel_id": get_val("id"),
            "title": title[:200],  # Limitar tamanho
            "status": get_val("status"),
            "assignee": get_val("assignee"),
            "priority": get_val("priority"),
            "area": get_val("area"),
            "created": get_val("created"),
            "due_date": get_val("due_date"),
            "description": get_val("description"),
            "category": get_val("category"),
            "complexity": get_val("complexity"),
            "sprint": get_val("sprint"),
            "row": row_idx,
        }
        tasks.append(task)
    
    print(f"  ✅ {len(tasks)} demandas encontradas no Excel")
    return tasks

def get_column_id_for_status(status, col_map):
    """Retorna o ID da coluna baseado no status"""
    mapped_col = STATUS_MAP.get(status, "01. Backlog")
    return col_map.get(mapped_col, list(col_map.values())[0])

def get_swimlane_for_priority(priority, sw_map):
    """Retorna o ID da swimlane baseado na prioridade"""
    if priority in ["Alta", "Crítica", "Urgente"]:
        key = "🔴 Alta Prioridade"
    elif priority in ["Média"]:
        key = "🟡 Média Prioridade"
    elif priority in ["Baixa"]:
        key = "🟢 Baixa Prioridade"
    else:
        key = "📋 Backlog Geral"
    
    # Fallback para primeira swimlane disponível
    if key not in sw_map:
        return list(sw_map.values())[0] if sw_map else 0
    return sw_map[key]

def get_priority_int(priority_str):
    return PRIORITY_MAP.get(priority_str, 2)

def get_color_for_priority(priority_str):
    p = get_priority_int(priority_str)
    return COLOR_MAP.get(p, "blue")

def delete_all_tasks_in_project(project_id):
    """Remove todas as tarefas de um projeto"""
    tasks = api("getAllTasks", {"project_id": project_id, "status_id": 1}) or []
    tasks_closed = api("getAllTasks", {"project_id": project_id, "status_id": 0}) or []
    all_tasks = tasks + tasks_closed
    
    deleted = 0
    for t in all_tasks:
        if api("removeTask", {"task_id": t["id"]}):
            deleted += 1
    return deleted

def main():
    print("=" * 60)
    print("🚀 SETUP INSANO DO KANBOARD - EBL SOLUÇÕES CORPORATIVAS")
    print("=" * 60)
    
    # 1. Verificar conexão
    me = api("getMe")
    if not me:
        print("❌ Falha na autenticação!")
        return
    print(f"\n✅ Conectado como: {me['name']} ({me['username']})")
    
    # 2. Listar projetos existentes
    projects = api("getAllProjects") or []
    print(f"\n📁 {len(projects)} projetos encontrados:")
    for p in projects:
        print(f"  ID={p['id']} | {p['name']}")
    
    # 3. Ler dados do Excel
    tasks_data = read_excel_data()
    
    # 4. Identificar o projeto principal (Fast Track Salesforce)
    sf_project = None
    for p in projects:
        if "SF" in p["name"] or "Fast Track" in p["name"] or "Salesforce" in p["name"]:
            sf_project = p
            break
    
    if not sf_project:
        print("❌ Projeto Fast Track Salesforce não encontrado!")
        return
    
    print(f"\n🎯 Projeto principal: {sf_project['name']} (ID={sf_project['id']})")
    
    # 5. Configurar TODOS os projetos com estrutura padronizada
    project_configs = {}
    
    for p in projects:
        pid = p["id"]
        pname = p["name"]
        print(f"\n{'='*50}")
        print(f"⚙️  Configurando: {pname}")
        
        col_map = setup_project_columns(pid, pname)
        sw_map = setup_project_swimlanes(pid, pname)
        cat_map = setup_project_categories(pid, pname)
        
        project_configs[pid] = {
            "name": pname,
            "columns": col_map,
            "swimlanes": sw_map,
            "categories": cat_map
        }
        print(f"  ✅ {pname} configurado!")
    
    # 6. Limpar tarefas duplicadas do projeto SF
    print(f"\n🧹 Limpando tarefas duplicadas do projeto {sf_project['name']}...")
    deleted = delete_all_tasks_in_project(sf_project["id"])
    print(f"  🗑️  {deleted} tarefas removidas")
    time.sleep(2)
    
    # 7. Recarregar configuração do projeto SF
    sf_config = project_configs[sf_project["id"]]
    col_map = sf_config["columns"]
    sw_map = sf_config["swimlanes"]
    cat_map = sf_config["categories"]
    
    print(f"\n  Colunas disponíveis: {list(col_map.keys())}")
    print(f"  Swimlanes disponíveis: {list(sw_map.keys())}")
    
    # 8. Importar todas as demandas do Excel
    print(f"\n📥 Importando {len(tasks_data)} demandas no projeto {sf_project['name']}...")
    
    imported = 0
    errors = 0
    
    for i, task in enumerate(tasks_data, 1):
        # Determinar coluna
        col_id = get_column_id_for_status(task["status"], col_map)
        
        # Determinar swimlane
        sw_id = get_swimlane_for_priority(task["priority"], sw_map)
        
        # Determinar prioridade numérica
        priority = get_priority_int(task["priority"])
        
        # Determinar cor
        color = get_color_for_priority(task["priority"])
        
        # Determinar categoria
        cat_id = 0
        for cat_name, cat_id_val in cat_map.items():
            if "Bug" in cat_name and "bug" in task.get("category","").lower():
                cat_id = cat_id_val
                break
            elif "Nova" in cat_name and "funcionalidade" in task.get("category","").lower():
                cat_id = cat_id_val
                break
        
        # Construir título com ID do Excel
        title = task["title"]
        if task["excel_id"] and task["excel_id"] != "None":
            title = f"[{task['excel_id']}] {task['title']}"
        title = title[:200]
        
        # Construir descrição rica
        desc_parts = []
        if task["description"] and task["description"] != "None":
            desc_parts.append(f"**Descrição:** {task['description']}")
        if task["area"] and task["area"] != "None":
            desc_parts.append(f"**Área/Solicitante:** {task['area']}")
        if task["complexity"] and task["complexity"] != "None":
            desc_parts.append(f"**Complexidade:** {task['complexity']}")
        if task["sprint"] and task["sprint"] != "None":
            desc_parts.append(f"**Sprint:** {task['sprint']}")
        if task["excel_id"] and task["excel_id"] != "None":
            desc_parts.append(f"**ID Cherwell:** {task['excel_id']}")
        if task["created"] and task["created"] != "None":
            desc_parts.append(f"**Data Abertura:** {task['created']}")
        
        description = "\n\n".join(desc_parts) if desc_parts else ""
        
        # Parâmetros da tarefa
        task_params = {
            "project_id": sf_project["id"],
            "title": title,
            "column_id": col_id,
            "swimlane_id": sw_id,
            "priority": priority,
            "color_id": color,
            "description": description,
        }
        
        # Adicionar data de vencimento se disponível
        if task["due_date"] and task["due_date"] not in ["None", ""]:
            try:
                # Tentar parsear a data
                date_str = str(task["due_date"])
                for fmt in ["%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"]:
                    try:
                        dt = datetime.strptime(date_str[:10], fmt[:8])
                        task_params["date_due"] = dt.strftime("%Y-%m-%d")
                        break
                    except:
                        pass
            except:
                pass
        
        # Adicionar categoria se disponível
        if cat_id:
            task_params["category_id"] = cat_id
        
        # Criar tarefa
        result = api("createTask", task_params)
        
        if result:
            imported += 1
            if i % 10 == 0:
                print(f"  📌 {i}/{len(tasks_data)} importadas... ({imported} OK, {errors} erros)")
        else:
            errors += 1
            print(f"  ❌ Erro na tarefa {i}: {title[:60]}...")
        
        # Pequena pausa para não sobrecarregar
        if i % 20 == 0:
            time.sleep(1)
    
    print(f"\n✅ Importação concluída!")
    print(f"  📌 Importadas com sucesso: {imported}")
    print(f"  ❌ Erros: {errors}")
    print(f"  📊 Total processado: {len(tasks_data)}")
    
    # 9. Verificar contagem final
    time.sleep(2)
    final_tasks = api("getAllTasks", {"project_id": sf_project["id"], "status_id": 1}) or []
    print(f"\n📊 VERIFICAÇÃO FINAL:")
    print(f"  Total de tarefas no projeto: {len(final_tasks)}")
    
    # Distribuição por coluna
    col_count = {}
    for t in final_tasks:
        col_name = "Desconhecida"
        for cname, cid in col_map.items():
            if cid == t.get("column_id"):
                col_name = cname
                break
        col_count[col_name] = col_count.get(col_name, 0) + 1
    
    print(f"\n  📊 Distribuição por coluna:")
    for col, count in sorted(col_count.items()):
        bar = "█" * min(count, 30)
        print(f"    {col:25s} | {bar} {count}")
    
    print("\n" + "="*60)
    print("🎉 SETUP CONCLUÍDO COM SUCESSO!")
    print("="*60)
    print(f"\n🌐 Acesse: http://kanboard.eblsolucoescorp.tec.br")
    print(f"👤 Usuário: admin")
    print(f"🔑 Senha: Senha@2026")
    print(f"🔑 Token API: a43a8785a4487979964cd7e12fc8c56bbb6ef7a6fa64bcb6c45fa1afc6ff")

if __name__ == "__main__":
    main()
