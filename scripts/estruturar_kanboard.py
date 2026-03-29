#!/usr/bin/env python3
"""
Script de Estruturação Completa do Kanboard - EBL Soluções Corp
Arquitetura para Gerente de Projetos e Diretoria
"""

import requests
import json
import openpyxl
from datetime import datetime
import time
import sys

# ============================================================
# CONFIGURAÇÃO
# ============================================================
BASE_URL = "http://kanboard.eblsolucoescorp.tec.br"
API_URL = f"{BASE_URL}/jsonrpc.php"
ADMIN_USER = "admin"
ADMIN_PASS = "446685558386787e2b1e0cebb0141257bebea7b2496bc6e960559c950692"
EXCEL_PATH = "/home/ubuntu/upload/Base_Fast_Tracking_Outubro.xlsx"

session = requests.Session()
session.auth = (ADMIN_USER, ADMIN_PASS)
session.headers.update({"Content-Type": "application/json"})

_req_id = 0

def api(method, params=None):
    global _req_id
    _req_id += 1
    payload = {"jsonrpc": "2.0", "method": method, "id": _req_id, "params": params or {}}
    try:
        r = session.post(API_URL, json=payload, timeout=30)
        data = r.json()
        if "error" in data:
            print(f"  [API ERROR] {method}: {data['error']}")
            return None
        return data.get("result")
    except Exception as e:
        print(f"  [EXCEPTION] {method}: {e}")
        return None

def log(msg, level="INFO"):
    icons = {"INFO": "ℹ️", "OK": "✅", "WARN": "⚠️", "ERR": "❌", "STEP": "🔷"}
    print(f"{icons.get(level,'•')} {msg}")

# ============================================================
# MAPEAMENTO DE COLUNAS (FASES) - SANITIZADO
# ============================================================
COLUNAS_WORKFLOW = [
    {"title": "01. Backlog",          "position": 1,  "color": "#95a5a6"},
    {"title": "02. Refinamento",      "position": 2,  "color": "#3498db"},
    {"title": "03. Priorizada",       "position": 3,  "color": "#2980b9"},
    {"title": "04. Análise",          "position": 4,  "color": "#f39c12"},
    {"title": "05. Estimativa",       "position": 5,  "color": "#e67e22"},
    {"title": "06. Aprovação",        "position": 6,  "color": "#e74c3c"},
    {"title": "07. Desenvolvimento",  "position": 7,  "color": "#9b59b6"},
    {"title": "08. Homologação",      "position": 8,  "color": "#1abc9c"},
    {"title": "09. Deploy",           "position": 9,  "color": "#27ae60"},
    {"title": "10. Implementado",     "position": 10, "color": "#2ecc71"},
    {"title": "11. Cancelado",        "position": 11, "color": "#7f8c8d"},
]

# Mapeamento de fases do Excel para colunas do Kanboard
FASE_MAP = {
    "backlog/sem priorização": "01. Backlog",
    "backlog": "01. Backlog",
    "sem priorização": "01. Backlog",
    "pendente": "01. Backlog",
    "refinamento": "02. Refinamento",
    " refinamento": "02. Refinamento",
    "priorizada": "03. Priorizada",
    "em análise": "04. Análise",
    "análise": "04. Análise",
    "em estimativa": "05. Estimativa",
    "estimativa": "05. Estimativa",
    "aguardando aprovação": "06. Aprovação",
    "em aprovação": "06. Aprovação",
    "aprovação": "06. Aprovação",
    "em desenvolvimento": "07. Desenvolvimento",
    "desenvolvimento": "07. Desenvolvimento",
    "em atendimento": "07. Desenvolvimento",
    "atendimento": "07. Desenvolvimento",
    "em homologação": "08. Homologação",
    "homologação": "08. Homologação",
    "homogação": "08. Homologação",
    "deploy": "09. Deploy",
    "em implementação": "09. Deploy",
    "implementação": "09. Deploy",
    "implementado": "10. Implementado",
    "concluído": "10. Implementado",
    "cancelado": "11. Cancelado",
}

# Mapeamento de prioridade
PRIORIDADE_MAP = {
    1: 3,  # P1 -> Urgente
    2: 2,  # P2 -> Alta
    3: 1,  # P3 -> Normal
    4: 0,  # P4 -> Baixa
    "1": 3, "2": 2, "3": 1, "4": 0,
}

# Cores por área
AREA_COLORS = {
    "Ar & Eletro":              "orange",
    "Automação":                "purple",
    "Bens de consumo":          "blue",
    "Cadastros":                "yellow",
    "Comercial":                "green",
    "Corporativo":              "grey",
    "Diretoria":                "red",
    "Financeiro":               "lime",
    "Fiscal":                   "teal",
    "Logística":                "cyan",
    "Marketing-Refrigeração":   "pink",
    "Programação de Materiais": "brown",
    "Refrigeração":             "sky",
    "TI":                       "navy",
    "Tecnologia da Informação": "navy",
}

# ============================================================
# PROJETOS (BOARDS)
# ============================================================
PROJETOS = [
    {
        "name": "[SF] Fast Track — Salesforce",
        "desc": "Gestão centralizada de todas as demandas do Fast Track Salesforce. Visão analítica para diretoria e gerência de projetos.",
        "identifier": "SF",
    },
    {
        "name": "[TI] Demandas de Tecnologia",
        "desc": "Demandas internas de TI, infraestrutura e sistemas.",
        "identifier": "TI",
    },
    {
        "name": "[COM] Demandas Comercial",
        "desc": "Demandas da área comercial, CRM e produtividade.",
        "identifier": "COM",
    },
    {
        "name": "[FIN] Demandas Financeiro",
        "desc": "Demandas do financeiro, fiscal e controladoria.",
        "identifier": "FIN",
    },
    {
        "name": "[OPS] Operações & Logística",
        "desc": "Demandas de operações, logística e supply chain.",
        "identifier": "OPS",
    },
]

# ============================================================
# SWIMLANES (RESPONSÁVEIS)
# ============================================================
SWIMLANES = [
    "Erick Almeida",
    "Marcio Souza",
    "Elder Rodrigues",
    "Felipe Nascimento",
    "Carlos Almeida",
]

# ============================================================
# CATEGORIAS (TIPOS DE DEMANDA)
# ============================================================
CATEGORIAS = [
    {"name": "Parametrização",          "color": "#3498db"},
    {"name": "Produtividade Comercial", "color": "#2ecc71"},
    {"name": "Produtividade TI",        "color": "#9b59b6"},
    {"name": "Incidente",               "color": "#e74c3c"},
    {"name": "Melhoria",                "color": "#f39c12"},
    {"name": "Novo Recurso",            "color": "#1abc9c"},
    {"name": "Integração",              "color": "#e67e22"},
]

# ============================================================
# FUNÇÕES DE SETUP
# ============================================================

def get_existing_projects():
    projects = api("getAllProjects") or []
    return {p["name"]: p for p in projects}

def get_project_columns(project_id):
    cols = api("getColumns", {"project_id": project_id}) or []
    return {c["title"]: c for c in cols}

def get_project_swimlanes(project_id):
    lanes = api("getAllSwimlanes", {"project_id": project_id}) or []
    return {s["name"]: s for s in lanes}

def get_project_categories(project_id):
    cats = api("getAllCategories", {"project_id": project_id}) or []
    return {c["name"]: c for c in cats}

def setup_project(proj_config):
    """Cria ou obtém projeto e configura colunas, swimlanes e categorias"""
    existing = get_existing_projects()
    
    if proj_config["name"] in existing:
        project = existing[proj_config["name"]]
        pid = project["id"]
        log(f"Projeto existente: {proj_config['name']} (ID: {pid})", "WARN")
    else:
        pid = api("createProject", {
            "name": proj_config["name"],
            "description": proj_config["desc"],
            "identifier": proj_config["identifier"],
        })
        if not pid:
            log(f"Falha ao criar projeto: {proj_config['name']}", "ERR")
            return None
        log(f"Projeto criado: {proj_config['name']} (ID: {pid})", "OK")
    
    # Configurar colunas
    setup_columns(pid)
    
    # Configurar swimlanes
    setup_swimlanes(pid)
    
    # Configurar categorias
    setup_categories(pid)
    
    return pid

def setup_columns(project_id):
    """Remove colunas antigas e cria as novas sanitizadas"""
    existing = get_project_columns(project_id)
    
    # Remover colunas antigas que não estão no novo layout
    new_titles = {c["title"] for c in COLUNAS_WORKFLOW}
    for title, col in existing.items():
        if title not in new_titles:
            result = api("removeColumn", {"column_id": col["id"]})
            log(f"  Coluna removida: '{title}'", "INFO")
    
    # Atualizar/criar colunas
    existing = get_project_columns(project_id)
    for col_config in COLUNAS_WORKFLOW:
        if col_config["title"] in existing:
            api("updateColumn", {
                "column_id": existing[col_config["title"]]["id"],
                "title": col_config["title"],
                "position": col_config["position"],
            })
        else:
            result = api("addColumn", {
                "project_id": project_id,
                "title": col_config["title"],
                "position": col_config["position"],
            })
            if result:
                log(f"  Coluna criada: '{col_config['title']}'", "OK")

def setup_swimlanes(project_id):
    """Configura swimlanes por responsável"""
    existing = get_project_swimlanes(project_id)
    
    for name in SWIMLANES:
        if name not in existing:
            result = api("addSwimlane", {
                "project_id": project_id,
                "name": name,
            })
            if result:
                log(f"  Swimlane criada: '{name}'", "OK")

def setup_categories(project_id):
    """Configura categorias de tipo de demanda"""
    existing = get_project_categories(project_id)
    
    for cat in CATEGORIAS:
        if cat["name"] not in existing:
            result = api("createCategory", {
                "project_id": project_id,
                "name": cat["name"],
            })
            if result:
                log(f"  Categoria criada: '{cat['name']}'", "OK")

# ============================================================
# IMPORTAÇÃO DE DEMANDAS
# ============================================================

def load_excel_data():
    """Carrega dados do Excel"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Status Report']
    
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    
    demandas = []
    for row in range(2, ws.max_row+1):
        vals = {headers[c-1]: ws.cell(row, c).value for c in range(1, ws.max_column+1)}
        
        # Pular linhas vazias
        if not vals.get('ID - Cherwell') and not vals.get('Tópico'):
            continue
        
        demandas.append(vals)
    
    return demandas

def sanitize_fase(fase_raw):
    """Normaliza fase para coluna do Kanboard"""
    if not fase_raw:
        return "01. Backlog"
    fase_lower = str(fase_raw).strip().lower()
    return FASE_MAP.get(fase_lower, "01. Backlog")

def sanitize_tipo(tipo_raw):
    """Normaliza tipo de demanda para categoria"""
    if not tipo_raw:
        return "Melhoria"
    tipo = str(tipo_raw).strip()
    # Mapear tipos do Excel para categorias
    tipo_map = {
        "Parametrização": "Parametrização",
        "Produtividade Comercial": "Produtividade Comercial",
        "Produtividade TI": "Produtividade TI",
        "Incidente ": "Incidente",
        "Incidente": "Incidente",
    }
    return tipo_map.get(tipo, "Melhoria")

def get_column_id(project_id, col_title):
    """Obtém ID de uma coluna pelo título"""
    cols = get_project_columns(project_id)
    col = cols.get(col_title)
    return col["id"] if col else None

def get_swimlane_id(project_id, name):
    """Obtém ID de swimlane pelo nome"""
    lanes = get_project_swimlanes(project_id)
    lane = lanes.get(name)
    return lane["id"] if lane else None

def get_category_id(project_id, name):
    """Obtém ID de categoria pelo nome"""
    cats = get_project_categories(project_id)
    cat = cats.get(name)
    return cat["id"] if cat else None

def get_existing_tasks(project_id):
    """Obtém tarefas existentes pelo título para evitar duplicatas"""
    tasks = api("getAllTasks", {"project_id": project_id, "status_id": 1}) or []
    return {t["title"]: t for t in tasks}

def format_date(dt):
    """Formata datetime para timestamp Unix"""
    if not dt:
        return None
    if isinstance(dt, datetime):
        return int(dt.timestamp())
    return None

def import_demandas(project_id, demandas):
    """Importa todas as demandas no projeto principal"""
    existing_tasks = get_existing_tasks(project_id)
    cols = get_project_columns(project_id)
    cats = get_project_categories(project_id)
    lanes = get_project_swimlanes(project_id)
    
    created = 0
    skipped = 0
    errors = 0
    
    for d in demandas:
        topico = d.get('Tópico') or d.get('Título') or "Sem título"
        topico = str(topico).strip()
        
        if not topico or topico == "None":
            continue
        
        # Verificar duplicata
        if topico in existing_tasks:
            skipped += 1
            continue
        
        # Mapear fase para coluna
        fase_raw = d.get('Fase Atual', '')
        col_title = sanitize_fase(fase_raw)
        col = cols.get(col_title)
        column_id = col["id"] if col else None
        
        # Mapear tipo para categoria
        tipo_raw = d.get('Tipo de Demanda', '')
        cat_name = sanitize_tipo(tipo_raw)
        cat = cats.get(cat_name)
        category_id = cat["id"] if cat else None
        
        # Mapear responsável para swimlane
        resp_raw = d.get('Responsavel pela demanda', '')
        resp = str(resp_raw).strip() if resp_raw else ''
        lane = lanes.get(resp)
        swimlane_id = lane["id"] if lane else None
        
        # Prioridade
        prio_raw = d.get('Prioridade area ', d.get('Prioridade', 3))
        priority = PRIORIDADE_MAP.get(prio_raw, 1)
        
        # Datas
        due_date = format_date(d.get('Go Live'))
        
        # ID Cherwell
        cherwell_id = d.get('ID - Cherwell', '')
        rdm = d.get('N°RDM', '')
        valtech = d.get('N° Valtech', '')
        horas_v = d.get('Horas estimadas Valtech', '')
        horas_e = d.get('Horas estimadas Elgin', '')
        area = d.get('Área Solicitante', '')
        requisitante = d.get('Requisitante', '')
        aprovado = d.get('Aprovado ?', '')
        obs = d.get('Obs:', '')
        
        # Montar descrição rica
        desc_parts = []
        if cherwell_id:
            desc_parts.append(f"**ID Cherwell:** {cherwell_id}")
        if rdm:
            desc_parts.append(f"**N° RDM:** {rdm}")
        if valtech:
            desc_parts.append(f"**N° Valtech:** {valtech}")
        if area:
            desc_parts.append(f"**Área Solicitante:** {area}")
        if requisitante:
            desc_parts.append(f"**Requisitante:** {requisitante}")
        if horas_v:
            desc_parts.append(f"**Horas Valtech:** {horas_v}h")
        if horas_e:
            desc_parts.append(f"**Horas Elgin:** {horas_e}h")
        if aprovado:
            desc_parts.append(f"**Aprovado:** {aprovado}")
        if obs:
            desc_parts.append(f"\n**Observações:** {obs}")
        
        description = "\n".join(desc_parts)
        
        # Cor por área
        color = AREA_COLORS.get(str(area).strip(), "blue") if area else "blue"
        
        # Criar tarefa
        task_params = {
            "project_id": project_id,
            "title": topico,
            "description": description,
            "priority": priority,
            "color_id": color,
        }
        
        if column_id:
            task_params["column_id"] = column_id
        if category_id:
            task_params["category_id"] = category_id
        if swimlane_id:
            task_params["swimlane_id"] = swimlane_id
        if due_date:
            task_params["date_due"] = due_date
        
        task_id = api("createTask", task_params)
        
        if task_id:
            created += 1
            # Adicionar referência externa como comentário
            ref_parts = []
            if cherwell_id:
                ref_parts.append(f"Cherwell: #{cherwell_id}")
            if d.get('Sequencia Demanda'):
                ref_parts.append(f"Seq: {d['Sequencia Demanda']}")
            
            if ref_parts:
                api("createComment", {
                    "task_id": task_id,
                    "user_id": 1,
                    "content": f"📋 Referências: {' | '.join(ref_parts)}\n📅 Fase original: {fase_raw}\n👤 Responsável: {resp}",
                })
            
            if created % 10 == 0:
                log(f"  {created} demandas importadas...", "INFO")
        else:
            errors += 1
        
        time.sleep(0.1)  # Rate limiting
    
    return created, skipped, errors

# ============================================================
# EXECUÇÃO PRINCIPAL
# ============================================================

def main():
    log("=" * 60, "STEP")
    log("ESTRUTURAÇÃO COMPLETA DO KANBOARD - EBL SOLUÇÕES CORP", "STEP")
    log("=" * 60, "STEP")
    
    # Verificar conexão
    me = api("getMe")
    if not me:
        log("Falha na autenticação! Verifique usuário e senha.", "ERR")
        sys.exit(1)
    log(f"Conectado como: {me.get('name', me.get('username'))}", "OK")
    
    # --------------------------------------------------------
    # FASE 1: Configurar projeto principal Fast Track
    # --------------------------------------------------------
    log("\n[FASE 1] Configurando projeto principal Fast Track...", "STEP")
    
    ft_config = PROJETOS[0]  # [SF] Fast Track
    pid_ft = setup_project(ft_config)
    
    if not pid_ft:
        log("Falha ao configurar projeto principal!", "ERR")
        sys.exit(1)
    
    log(f"Projeto principal configurado (ID: {pid_ft})", "OK")
    
    # --------------------------------------------------------
    # FASE 2: Configurar demais projetos
    # --------------------------------------------------------
    log("\n[FASE 2] Configurando demais projetos...", "STEP")
    
    project_ids = {ft_config["name"]: pid_ft}
    
    for proj in PROJETOS[1:]:
        pid = setup_project(proj)
        if pid:
            project_ids[proj["name"]] = pid
            log(f"Projeto configurado: {proj['name']} (ID: {pid})", "OK")
    
    # --------------------------------------------------------
    # FASE 3: Importar demandas do Excel
    # --------------------------------------------------------
    log("\n[FASE 3] Importando demandas do Excel...", "STEP")
    
    demandas = load_excel_data()
    log(f"Total de demandas no Excel: {len(demandas)}", "INFO")
    
    created, skipped, errors = import_demandas(pid_ft, demandas)
    
    log(f"Importação concluída:", "OK")
    log(f"  ✅ Criadas: {created}", "INFO")
    log(f"  ⏭️  Ignoradas (duplicatas): {skipped}", "INFO")
    log(f"  ❌ Erros: {errors}", "INFO")
    
    # --------------------------------------------------------
    # FASE 4: Relatório final
    # --------------------------------------------------------
    log("\n[FASE 4] Gerando relatório...", "STEP")
    
    for proj_name, pid in project_ids.items():
        tasks = api("getAllTasks", {"project_id": pid, "status_id": 1}) or []
        log(f"  {proj_name}: {len(tasks)} tarefas", "INFO")
    
    log("\n" + "=" * 60, "STEP")
    log("ESTRUTURAÇÃO CONCLUÍDA COM SUCESSO!", "OK")
    log(f"Acesse: {BASE_URL}", "OK")
    log("=" * 60, "STEP")
    
    return {
        "projects": project_ids,
        "demandas_criadas": created,
        "demandas_ignoradas": skipped,
        "erros": errors,
    }

if __name__ == "__main__":
    result = main()
    print(f"\nResultado: {json.dumps(result, indent=2)}")
